from django.shortcuts import render, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods
from .models import Empleado, RegistroAsistencia
from .services import ReporteService
from .controllers import (
    procesar_registro_por_empleado,
    procesar_registro_manual,
    crear_contexto_exito,
    buscar_empleado_por_qr_api,
    identificar_por_fingerprint_api,
    vincular_fingerprint_api,
    desvincular_fingerprint_api,
)
import openpyxl
import string
from datetime import date, timedelta
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import json


def es_staff(user):
    if not hasattr(user, 'is_authenticated'):
        return False
    return bool(user.is_authenticated and getattr(user, 'is_staff', False))


def _sanitize_sheet_title(title):
    invalid_chars = set('[]:*?/\\')
    clean_title = ''.join('_' if c in invalid_chars else c for c in title)
    return clean_title[:31]


def _sanitize_table_name(name):
    valid_chars = []
    for c in name:
        if c in string.ascii_letters or c in string.digits or c == '_':
            valid_chars.append(c)
        else:
            valid_chars.append('_')
    clean_name = ''.join(valid_chars)
    if not clean_name or not (clean_name[0].isalpha() or clean_name[0] == '_'):
        clean_name = 'T' + clean_name
    return clean_name[:31]


def _make_unique_name(base, existing, max_length=31):
    if base not in existing:
        return base

    suffix = 1
    while True:
        suffix_str = f"_{suffix}"
        truncated_base = base[:max_length - len(suffix_str)]
        candidate = f"{truncated_base}{suffix_str}"
        if candidate not in existing:
            return candidate
        suffix += 1


def _make_unique_sheet_title(title, existing_titles, max_length=31):
    title = _sanitize_sheet_title(title)
    return _make_unique_name(title, existing_titles, max_length)


@user_passes_test(es_staff)
def pagina_descarga_excel(request):
    return render(request, 'pagina_descarga_excel.html')


@user_passes_test(es_staff)
def exportar_resumen_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Diario"

    encabezados = [
        "Empleado", "Fecha", "Tiempo de Almuerzo",
        "Horas por Comisión", "Horas por Permiso (Otros)",
        "Horas Trabajadas Totales"
    ]
    ws.append(encabezados)

    datos_diarios = ReporteService.obtener_datos_resumen()

    for (id_empleado, fecha), data in datos_diarios.items():
        empleado = data["empleado"]
        horas = ReporteService.calcular_horas_empleado(data)
        ws.append([
            empleado.nombre_completo,
            fecha.strftime("%Y-%m-%d"),
            horas['almuerzo'],
            horas['comision'],
            horas['permiso'],
            horas['trabajadas']
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F1D87")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    if ws.max_row > 1:
        tabla = Table(
            displayName="ResumenAsistencia",
            ref=f"A1:F{ws.max_row}"
        )
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        tabla.tableStyleInfo = style
        ws.add_table(tabla)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=resumen_asistencia.xlsx'
    wb.save(response)
    return response


@user_passes_test(es_staff)
def exportar_asistencia_excel(request):
    fecha_inicio = date(2026, 6, 29)
    fecha_fin = date(2026, 7, 31)
    semanas = ReporteService.obtener_semanas_en_rango(fecha_inicio, fecha_fin)
    registros_semanales = ReporteService.obtener_registros_semanales(fecha_inicio, fecha_fin)

    wb = openpyxl.Workbook()
    empleados = Empleado.objects.order_by('apellidos', 'nombres')
    existing_titles = set()

    for index, empleado in enumerate(empleados):
        if index == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()

        sheet_title = _make_unique_sheet_title(empleado.nombre_completo, existing_titles)
        existing_titles.add(sheet_title)
        ws.title = sheet_title

        ws.append(["REGISTRO DE ASISTENCIA"] + [""] * 8)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        ws.append(["Nombre completo:", empleado.nombre_completo] + [""] * 7)
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=9)
        ws.append([])

        encabezados = [
            "Fecha", "Entrada", "Inicio Almuerzo", "Fin Almuerzo", "Salida",
            "Entrada por comisión", "Salida por comisión", "Entrada por otros", "Salida por otros"
        ]

        registros_por_semana = registros_semanales.get(empleado.id_empleado, {})

        for semana_num, semana_inicio in enumerate(semanas, start=1):
            semana_fin = semana_inicio + timedelta(days=4)
            week_title = f"Semana {semana_num} ({semana_inicio.strftime('%d/%m/%Y')} - {semana_fin.strftime('%d/%m/%Y')})"
            ws.append([week_title] + [""] * 8)
            ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=9)
            ws.append(encabezados)

            for dia_offset in range(5):
                fecha_dia = semana_inicio + timedelta(days=dia_offset)
                fecha_str = fecha_dia.strftime('%d/%m/%Y')
                if fecha_dia < fecha_inicio or fecha_dia > fecha_fin:
                    ws.append([fecha_str] + [''] * 8)
                    continue

                datos_dia = registros_por_semana.get(semana_inicio, {}).get(fecha_dia)
                if not datos_dia:
                    ws.append([fecha_str] + [''] * 8)
                    continue

                ws.append([
                    fecha_str,
                    ReporteService.obtener_hora_tipo(datos_dia, 'Entrada'),
                    ReporteService.obtener_hora_tipo(datos_dia, 'Inicio Almuerzo'),
                    ReporteService.obtener_hora_tipo(datos_dia, 'Fin Almuerzo'),
                    ReporteService.obtener_hora_tipo(datos_dia, 'Salida'),
                    ReporteService.obtener_hora_tipo(datos_dia, 'Entrada por comisión'),
                    ReporteService.obtener_hora_tipo(datos_dia, 'Salida por comisión'),
                    ReporteService.obtener_hora_tipo(datos_dia, 'Entrada por otros'),
                    ReporteService.obtener_hora_tipo(datos_dia, 'Salida por otros'),
                ])

            ws.append([])

        from openpyxl.utils import get_column_letter

        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            if max_length == 0:
                max_length = 0
            first_cell = col[0]
            column_index = getattr(first_cell, 'column', None)
            if column_index is None:
                continue
            column_letter = get_column_letter(column_index)
            ws.column_dimensions[column_letter].width = max_length + 2

        header_font = Font(bold=True, color="FFFFFF")
        week_title_fill = PatternFill("solid", fgColor="0F1D87")
        column_header_fill = PatternFill("solid", fgColor="4F81BD")
        title_fill = PatternFill("solid", fgColor="0F1D87")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9):
            first_cell = row[0].value
            if first_cell == 'REGISTRO DE ASISTENCIA':
                row[0].font = Font(bold=True, color='FFFFFF', size=12)
                row[0].fill = title_fill
            if first_cell == 'Nombre completo:':
                row[0].font = Font(bold=True)
            if first_cell and str(first_cell).startswith('Semana '):
                row[0].font = Font(bold=True, color='FFFFFF')
                row[0].fill = week_title_fill
            if first_cell == 'Fecha':
                for cell in row:
                    cell.font = header_font
                    cell.fill = column_header_fill
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9):
            for cell in row:
                if cell.border is None:
                    cell.border = thin_border

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=asistencia_semanal_por_empleado.xlsx'
    wb.save(response)
    return response


def pagina_principal(request):
    return render(request, 'pagina_principal.html')


def escanear_qr(request):
    return render(request, 'escanear_qr.html')


@ensure_csrf_cookie
def identificar_dispositivo(request):
    empleados = Empleado.objects.order_by('apellidos', 'nombres')
    return render(request, 'identificar.html', {'empleados': empleados})


def registrar_asistencia_qr(request, codigo_qr):
    empleado = Empleado.buscar_por_codigo_qr(codigo_qr)
    if not empleado:
        messages.error(request, 'Código QR no válido o empleado no encontrado.')
        return render(request, 'error_qr.html')

    resultado = procesar_registro_por_empleado(request, empleado)
    if resultado['is_post']:
        if resultado['success']:
            messages.success(request, resultado['message'])
            return render(request, 'asistencia_exitosa.html', crear_contexto_exito(resultado['registro']))
        messages.error(request, resultado['message'])

    return render(request, 'formulario_qr.html', resultado['context'])


def registrar_asistencia_auto(request, empleado_id):
    empleado = get_object_or_404(Empleado, id_empleado=empleado_id)
    resultado = procesar_registro_por_empleado(request, empleado)
    if resultado['is_post']:
        if resultado['success']:
            messages.success(request, resultado['message'])
            return render(request, 'asistencia_exitosa.html', crear_contexto_exito(resultado['registro']))
        messages.error(request, resultado['message'])

    return render(request, 'formulario_qr.html', resultado['context'])


@require_http_methods(["POST", "OPTIONS"])
def api_buscar_empleado_qr(request):
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})

    try:
        data = json.loads(request.body)
        resultado, status_code = buscar_empleado_por_qr_api(data)
        return JsonResponse(resultado, status=status_code)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_identificar_por_fingerprint(request):
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})

    try:
        data = json.loads(request.body)
        resultado, status_code = identificar_por_fingerprint_api(data)
        return JsonResponse(resultado, status=status_code)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_vincular_fingerprint(request):
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})

    try:
        data = json.loads(request.body)
        resultado, status_code = vincular_fingerprint_api(data)
        return JsonResponse(resultado, status=status_code)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


@require_http_methods(["POST", "OPTIONS"])
def api_desvincular_fingerprint(request):
    if request.method == 'OPTIONS':
        return JsonResponse({'success': True})

    try:
        data = json.loads(request.body)
        resultado, status_code = desvincular_fingerprint_api(data)
        return JsonResponse(resultado, status=status_code)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error del servidor: {str(e)}'}, status=500)


def registrar_asistencia(request):
    resultado = procesar_registro_manual(request)
    if resultado['is_post']:
        if resultado['success']:
            messages.success(request, resultado['message'])
            return render(request, 'asistencia_exitosa.html', crear_contexto_exito(resultado['registro']))
        messages.error(request, resultado['message'])

    return render(request, 'formulario.html', resultado['context'])
